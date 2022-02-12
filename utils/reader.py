from typing import List

# from keras.datasets import mnist
from openpyxl import load_workbook


class Reader:

    def __init__(self):
        pass
        # (train_X, train_y), (test_X, test_y) = mnist.load_data()
        # self.train_X = train_X
        # self.train_y = train_y
        # self.test_X = test_X
        # self.test_y = test_y

    @classmethod
    def read_train_data(cls):
        X = []
        table = load_workbook('number_train.xlsx')
        table_sheet_names = table.sheetnames
        for sheet in range(len(table.sheetnames)):
            X.append([])
            for i in range(5):
                for j in range(7):
                    X[len(X) - 1].append(
                        table.get_sheet_by_name(str(table_sheet_names[sheet])).cell(row=j + 1, column=i + 1).value)
        return X

    @classmethod
    def read_test_data(cls):
        X = []
        table = load_workbook('number_data.xlsx')
        table_sheet = table.get_sheet_by_name("0")
        for i in range(5):
            for j in range(7):
                X.append(table_sheet.cell(row=j + 1, column=i + 1).value)
        return X

    @staticmethod
    def get_target_data(number: int) -> List[float]:
        table = load_workbook('number_target.xlsx')
        target = []
        for i in range(10):
            target.append(table.get_sheet_by_name(str(1)).cell(row=i + 1, column=number + 1).value)
        return target
